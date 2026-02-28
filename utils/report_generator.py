"""
utils/report_generator.py — Generates downloadable reports

Outputs:
- Excel report: full analysis workbook per student
- CSV batch export: institution-wide batch results
- JSON export: raw profile + scores for API integration
"""
from __future__ import annotations
import io
import json
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List


# ── Style helpers ─────────────────────────────────────────────────────────
_thin = Side(style="thin", color="1C2740")
_bdr  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _hc(ws, row, col, value, bg="0D1421", fg="00D4FF", bold=True, size=10, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, color=fg, size=size)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    c.border = _bdr
    return c

def _dc(ws, row, col, value="", fg="E2E8F0", bg=None, bold=False):
    bg = bg or ("0F1826" if row % 2 == 0 else "0D1421")
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", color=fg, bold=bold, size=10)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = _bdr
    return c


# ── Student Excel Report ───────────────────────────────────────────────────

def generate_student_excel(
    profile: dict,
    career_results: list,
    swot: dict,
    benchmark: dict,
    cluster: dict,
) -> bytes:
    """
    Generate a full per-student Excel analysis report.
    Returns bytes for st.download_button.
    """
    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.sheet_view.showGridLines = False

    # Title
    ws1.merge_cells("A1:F1")
    t = ws1["A1"]
    t.value = f"⚡ SPECTRA Analysis Report — {profile.get('name', 'Student')}"
    t.font = Font(name="Arial", bold=True, color="00D4FF", size=14)
    t.fill = PatternFill("solid", fgColor="0D1421")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 34

    # Generated date
    ws1.merge_cells("A2:F2")
    d = ws1["A2"]
    d.value = f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}   |   IIT Techkriti 2026 · SPECTRA v2.0"
    d.font = Font(name="Arial", italic=True, color="4A5A7A", size=9)
    d.fill = PatternFill("solid", fgColor="080C14")
    d.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 18

    # Profile header row
    _hc(ws1, 4, 1, "Field",          bg="141D2E", size=9)
    _hc(ws1, 4, 2, "Value",          bg="141D2E", size=9)
    _hc(ws1, 4, 3, "Field",          bg="141D2E", size=9)
    _hc(ws1, 4, 4, "Value",          bg="141D2E", size=9)
    ws1.row_dimensions[4].height = 26

    profile_fields = [
        ("Name",            profile.get("name", "—"),             "College",        profile.get("college", "—")),
        ("Student ID",      profile.get("student_id", "—"),       "Branch",         profile.get("branch", "—")),
        ("Semester",        profile.get("semester", "—"),          "Year",           profile.get("year", "—")),
        ("CGPA",            f"{profile.get('cgpa', 0)}/10",        "Career Goal",    profile.get("career_goal", "—")),
        ("Effort Level",    f"{profile.get('effort', 0)}/5",       "Target",         profile.get("timeline", "—")),
        ("Academic Trend",  profile.get("academic_trend", "—"),    "Risk Level",     profile.get("risk_level", "—")),
    ]
    for r, (f1, v1, f2, v2) in enumerate(profile_fields, start=5):
        ws1.row_dimensions[r].height = 22
        _hc(ws1, r, 1, f1, bg="0D1421", fg="7A90B0", bold=False, size=9)
        _dc(ws1, r, 2, v1, fg="E2E8F0")
        _hc(ws1, r, 3, f2, bg="0D1421", fg="7A90B0", bold=False, size=9)
        _dc(ws1, r, 4, v2, fg="E2E8F0")

    # Score cards
    _hc(ws1, 12, 1, "Intelligence Score", bg="141D2E")
    _hc(ws1, 12, 2, "Career Fit %",       bg="141D2E")
    _hc(ws1, 12, 3, "CGPA Percentile",    bg="141D2E")
    _hc(ws1, 12, 4, "Learning Profile",   bg="141D2E")
    ws1.row_dimensions[12].height = 26

    top_fit = career_results[0]["fit"] if career_results else 0
    _hc(ws1, 13, 1, f"{profile.get('intelligence_score', 0)}/100",
        bg="0A1628", fg="00E887", size=14)
    _hc(ws1, 13, 2, f"{top_fit}%",  bg="0A1628", fg="FFB800", size=14)
    _hc(ws1, 13, 3, f"{benchmark.get('cgpa_percentile', '—')}th",
        bg="0A1628", fg="00D4FF", size=14)
    _hc(ws1, 13, 4, cluster.get("name", "—"), bg="0A1628", fg="A78BFA", size=11)
    ws1.row_dimensions[13].height = 34

    for col, width in zip("ABCDEF", [20, 22, 20, 22, 16, 16]):
        ws1.column_dimensions[col].width = width

    # ── Sheet 2: Career Analysis ──────────────────────────────────────
    ws2 = wb.create_sheet("Career Analysis")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:G1")
    t2 = ws2["A1"]
    t2.value = f"🎯 Career Fit Analysis — {profile.get('name', 'Student')}"
    t2.font = Font(name="Arial", bold=True, color="00D4FF", size=13)
    t2.fill = PatternFill("solid", fgColor="0D1421")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    headers = ["Rank", "Career", "Overall Fit %", "ML Score", "Formula Score", "Market Demand", "Est. Salary"]
    for ci, h in enumerate(headers, 1):
        _hc(ws2, 3, ci, h, bg="141D2E", size=9)
        ws2.row_dimensions[3].height = 26

    for ri, c in enumerate(career_results[:8], start=4):
        ws2.row_dimensions[ri].height = 22
        rank_color = ["FFD700", "C0C0C0", "CD7F32"][ri-4] if ri <= 6 else "4A5A7A"
        _hc(ws2, ri, 1, f"#{ri-3}", bg="0D1421", fg=rank_color, bold=True)
        _dc(ws2, ri, 2, c.get("title", "—"), fg="E2E8F0", bold=(ri==4))
        fit_fg = "00E887" if c["fit"] >= 75 else ("FFB800" if c["fit"] >= 55 else "FF4D6A")
        _hc(ws2, ri, 3, f"{c['fit']}%", bg="0D1421", fg=fit_fg, bold=True)
        _dc(ws2, ri, 4, f"{c.get('ml_score', '—')}%")
        _dc(ws2, ri, 5, f"{c.get('formula_score', '—')}%")
        _dc(ws2, ri, 6, c.get("demand", "—"))
        _dc(ws2, ri, 7, c.get("salary", "—"))

    for col, width in zip("ABCDEFG", [6, 22, 14, 12, 14, 14, 18]):
        ws2.column_dimensions[get_column_letter(col if isinstance(col, int) else ord(col)-64)].width = width

    # ── Sheet 3: Skills & SWOT ────────────────────────────────────────
    ws3 = wb.create_sheet("Skills & SWOT")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:F1")
    t3 = ws3["A1"]
    t3.value = f"📊 Skills Profile & SWOT Analysis — {profile.get('name', 'Student')}"
    t3.font = Font(name="Arial", bold=True, color="00D4FF", size=13)
    t3.fill = PatternFill("solid", fgColor="0D1421")
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 30

    # Skills
    _hc(ws3, 3, 1, "Skill",   bg="141D2E", size=9)
    _hc(ws3, 3, 2, "Score",   bg="141D2E", size=9)
    _hc(ws3, 3, 3, "Rating",  bg="141D2E", size=9)
    ws3.row_dimensions[3].height = 24

    skills = profile.get("skills", {})
    skill_items = [
        ("Technical",     skills.get("technical", 0)),
        ("Analytical",    skills.get("analytical", 0)),
        ("Creative",      skills.get("creative", 0)),
        ("Communication", skills.get("communication", 0)),
        ("Leadership",    skills.get("leadership", 0)),
        ("Consistency",   skills.get("consistency", 0)),
    ]
    for ri, (name, val) in enumerate(skill_items, start=4):
        ws3.row_dimensions[ri].height = 22
        _dc(ws3, ri, 1, name, fg="E2E8F0")
        fg = "00E887" if val >= 75 else ("FFB800" if val >= 50 else "FF4D6A")
        _hc(ws3, ri, 2, f"{val}/100", bg="0D1421", fg=fg)
        rating = "Excellent" if val >= 80 else ("Good" if val >= 65 else ("Fair" if val >= 50 else "Needs Work"))
        _dc(ws3, ri, 3, rating, fg=fg)

    # SWOT sections
    swot_sections = [
        (12, "💪 STRENGTHS",    swot.get("strengths", []),    "00E887"),
        (19, "⚠️ WEAKNESSES",   swot.get("weaknesses", []),   "FF4D6A"),
        (26, "🚀 OPPORTUNITIES",swot.get("opportunities", []),"FFB800"),
        (33, "🛡️ THREATS",      swot.get("threats", []),      "A78BFA"),
    ]
    for start_row, label, items, color in swot_sections:
        ws3.merge_cells(f"A{start_row}:F{start_row}")
        sh = ws3.cell(row=start_row, column=1, value=f"  {label}")
        sh.font = Font(name="Arial", bold=True, color=color, size=10)
        sh.fill = PatternFill("solid", fgColor="141D2E")
        sh.alignment = Alignment(horizontal="left", vertical="center")
        sh.border = _bdr
        ws3.row_dimensions[start_row].height = 24

        for i, item in enumerate(items[:5], start=1):
            ws3.merge_cells(f"A{start_row+i}:F{start_row+i}")
            ic = ws3.cell(row=start_row+i, column=1, value=f"  • {item}")
            ic.font = Font(name="Arial", color="E2E8F0", size=9)
            ic.fill = PatternFill("solid", fgColor="0D1421" if i % 2 else "0F1826")
            ic.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ic.border = _bdr
            ws3.row_dimensions[start_row+i].height = 30

    ws3.column_dimensions["A"].width = 80

    # ── Sheet 4: Benchmark ────────────────────────────────────────────
    ws4 = wb.create_sheet("Peer Benchmark")
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells("A1:D1")
    t4 = ws4["A1"]
    t4.value = f"📈 Peer Benchmarking — {profile.get('name', 'Student')}"
    t4.font = Font(name="Arial", bold=True, color="00D4FF", size=13)
    t4.fill = PatternFill("solid", fgColor="0D1421")
    t4.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 30

    _hc(ws4, 3, 1, "Metric",         bg="141D2E", size=9)
    _hc(ws4, 3, 2, "Your Value",     bg="141D2E", size=9)
    _hc(ws4, 3, 3, "Cohort Average", bg="141D2E", size=9)
    _hc(ws4, 3, 4, "Percentile",     bg="141D2E", size=9)
    ws4.row_dimensions[3].height = 26

    bm_rows = [
        ("CGPA",              f"{profile.get('cgpa',0)}/10",
         f"{benchmark.get('cgpa_cohort_avg','—')}/10",
         f"{benchmark.get('cgpa_percentile','—')}th"),
        ("Intelligence Score",f"{profile.get('intelligence_score',0)}/100",
         f"{benchmark.get('intel_cohort_avg','—')}/100",
         f"{benchmark.get('intel_percentile','—')}th"),
        ("Technical Skill",   f"{skills.get('technical',0)}/100",
         "—", f"{benchmark.get('skill_percentiles',{}).get('technical','—')}th"),
        ("Analytical Skill",  f"{skills.get('analytical',0)}/100",
         "—", f"{benchmark.get('skill_percentiles',{}).get('analytical','—')}th"),
        ("Peer Group Rank",   f"#{benchmark.get('peer_cgpa_rank','—')}",
         f"{benchmark.get('peer_count','—')} peers", "—"),
    ]
    for ri, (metric, val, avg, pct) in enumerate(bm_rows, start=4):
        ws4.row_dimensions[ri].height = 22
        _dc(ws4, ri, 1, metric, fg="E2E8F0", bold=True)
        _dc(ws4, ri, 2, val,   fg="FFB800")
        _dc(ws4, ri, 3, avg,   fg="7A90B0")
        pct_fg = "00E887" if str(pct).replace("th","").isdigit() and int(str(pct).replace("th","")) >= 75 else "FFB800"
        _dc(ws4, ri, 4, pct,   fg=pct_fg)

    for col, w in zip("ABCD", [28, 18, 18, 14]):
        ws4.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Batch CSV Export ───────────────────────────────────────────────────────

def generate_batch_csv(results_df: pd.DataFrame) -> bytes:
    """Export batch ML results to CSV."""
    return results_df.to_csv(index=False).encode()


# ── JSON Profile Export ────────────────────────────────────────────────────

def generate_profile_json(profile: dict, career_results: list, swot: dict) -> bytes:
    """Export full profile + analysis as JSON (for API integration)."""
    export = {
        "meta": {
            "generated_at": datetime.now().isoformat(),
            "system":       "SPECTRA v2.0",
            "event":        "IIT Techkriti 2026",
        },
        "profile":         profile,
        "career_analysis": career_results[:5],
        "swot":            swot,
    }
    return json.dumps(export, indent=2, default=str).encode()
